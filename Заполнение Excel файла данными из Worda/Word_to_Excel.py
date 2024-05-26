from docx import Document
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

def read_word_table(file_path, fio_column, teacher_column):
    doc = Document(file_path)
    data = {}

    for table in doc.tables:
        # Поиск индексов нужных столбцов
        fio_col_idx = None
        teacher_col_idx = None
        for i, cell in enumerate(table.rows[0].cells):
            if cell.text.strip() == fio_column:
                fio_col_idx = i
            elif cell.text.strip() == teacher_column:
                teacher_col_idx = i
            if fio_col_idx is not None and teacher_col_idx is not None:
                break
                
        if fio_col_idx is None or teacher_col_idx is None:
            continue  # Пропустить таблицы, не содержащие нужных столбцов
        
        for row in table.rows[1:]:
            fio = row.cells[fio_col_idx].text.strip()
            teacher = row.cells[teacher_col_idx].text.strip()
            data[fio] = teacher
            
    return data

def update_excel_with_data(excel_file_path, word_data, fio_column, teacher_column):
    try:
        # Открытие существующего Excel файла
        book = load_workbook(excel_file_path)
    except Exception as e:
        print(f"Ошибка при открытии файла Excel: {e}")
        return

    for sheet_name in book.sheetnames:
        sheet = book[sheet_name]
        df = pd.DataFrame(sheet.values)
        # Поиск индексов нужных столбцов
        try:
            fio_col_idx = df.iloc[0].tolist().index(fio_column)
            teacher_col_idx = df.iloc[0].tolist().index(teacher_column)
        except ValueError:
            print(f"Столбцы с именами '{fio_column}' и/или '{teacher_column}' не найдены на листе '{sheet_name}'")
            continue
        
        # Обновление данных в соответствующих столбцах
        for i in range(1, len(df)):
            fio = df.iat[i, fio_col_idx]
            if fio in word_data:
                df.iat[i, teacher_col_idx] = word_data[fio]

        # Очистка листа перед записью
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row):
            for cell in row:
                cell.value = None
        
        # Запись обновленного DataFrame обратно на лист
        for r_idx, row in df.iterrows():
            for c_idx, value in enumerate(row):
                sheet.cell(row=r_idx + 1, column=c_idx + 1, value=value)

    # Сохранение изменений в Excel файл
    try:
        book.save(excel_file_path)
    except Exception as e:
        print(f"Ошибка при сохранении файла Excel: {e}")

# Путь к файлам и имена столбцов
word_file_path = r'C:\Users\User\Desktop\test\Таблица участников.docx'
excel_file_path = r'C:\Users\User\Desktop\test\Сводная.xlsx'
fio_column = 'ФИО'
teacher_column = 'Преподаватель'

# Чтение данных из Word документа
word_data = read_word_table(word_file_path, fio_column, teacher_column)

# Обновление Excel файла
update_excel_with_data(excel_file_path, word_data, fio_column, teacher_column)
